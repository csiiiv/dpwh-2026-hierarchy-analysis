#!/usr/bin/env python3
"""
Phase 1 Only: Build row number hierarchy from Excel formulas.

Output: Row-based hierarchy (no labels or amounts)
Purpose: Fast formula extraction to build parent-child relationships
"""

import re
import json
import time
from pathlib import Path
from collections import defaultdict

def parse_formula_references(formula: str, target_column: str) -> list:
    """
    Parse formula to extract child row references.
    Handles:
    - Individual references: =K7+K14
    - SUM ranges: =SUM(K39:K55) -> includes ALL rows 39-55
    """
    if not formula or not formula.startswith('='):
        return []
    
    formula = formula[1:]  # Remove = sign
    
    # Check for SUM range
    sum_pattern = rf"SUM\({target_column}(\d+):{target_column}(\d+)\)"
    sum_match = re.search(sum_pattern, formula, re.IGNORECASE)
    
    if sum_match:
        return list(range(int(sum_match.group(1)), int(sum_match.group(2)) + 1))
    
    # Individual cell references
    cell_pattern = f"{target_column}\\d+"
    cell_refs = re.findall(cell_pattern, formula)
    row_numbers = [int(''.join(filter(str.isdigit, ref))) for ref in cell_refs]
    return sorted(set(row_numbers))

def build_formula_hierarchy(xlsx_path: Path, column: str = 'K') -> dict:
    """
    Build row-based hierarchy from formulas.
    
    Returns:
        Dictionary with row relationships (no labels or amounts)
    """
    print("="*70)
    print("PHASE 1: BUILDING ROW HIERARCHY FROM FORMULAS")
    print("="*70)
    
    from openpyxl import load_workbook
    
    start_time = time.time()
    
    wb = load_workbook(xlsx_path, data_only=False, read_only=True)
    sheet = wb.active
    
    col_idx = ord(column) - ord('A') + 1
    
    formula_cells = []
    parent_to_children = defaultdict(list)
    child_to_parents = defaultdict(list)
    
    max_row_limit = sheet.max_row
    print(f"Scanning rows 1-{max_row_limit} in column {column}...")
    
    for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, max_row=max_row_limit):
        for cell in row:
            if cell and cell.data_type == 'f':  # Formula cell
                formula = cell.value
                refs = parse_formula_references(formula, column)
                
                formula_cells.append({
                    'row': cell.row,
                    'formula': formula,
                    'child_rows': refs,
                    'num_children': len(refs)
                })
                
                for child_row in refs:
                    parent_to_children[cell.row].append(child_row)
                    child_to_parents[child_row].append(cell.row)
    
    wb.close()
    
    all_rows_involved = set(parent_to_children.keys())
    for child_rows in parent_to_children.values():
        all_rows_involved.update(child_rows)
    
    root_rows = [row for row in parent_to_children.keys() 
                  if row not in child_to_parents]
    
    hierarchy_tree = build_tree_from_relationships(
        parent_to_children, root_rows
    )
    
    elapsed = time.time() - start_time
    
    print(f"\n✓ Phase 1 complete in {elapsed:.2f}s")
    print(f"  Formula cells:  {len(formula_cells)}")
    print(f"  Parent nodes:   {len(parent_to_children)}")
    print(f"  Child nodes:     {len(all_rows_involved) - len(parent_to_children)}")
    print(f"  Total rows:     {len(all_rows_involved)}")
    print(f"  Root nodes:     {len(root_rows)}")
    
    return {
        'formula_cells': formula_cells,
        'parent_to_children': dict(parent_to_children),
        'child_to_parents': dict(child_to_parents),
        'all_rows_involved': list(all_rows_involved),
        'root_rows': sorted(root_rows),
        'hierarchy_tree': hierarchy_tree,
        'stats': {
            'elapsed_time': elapsed,
            'num_formulas': len(formula_cells),
            'num_parents': len(parent_to_children),
            'num_children': len(all_rows_involved) - len(parent_to_children),
            'num_roots': len(root_rows)
        }
    }

def build_tree_from_relationships(parent_to_children: dict, root_rows: list) -> list:
    """Build tree structure from parent-child relationships."""
    tree = []
    
    def build_node(row: int) -> dict:
        children_rows = parent_to_children.get(row, [])
        
        node = {
            'row': row,
            'children': []
        }
        
        for child_row in children_rows:
            child_node = build_node(child_row)
            node['children'].append(child_node)
        
        return node
    
    for root_row in sorted(root_rows):
        tree.append(build_node(root_row))
    
    return tree

def save_row_hierarchy(hierarchy: dict, output_path: Path):
    """Save row hierarchy to JSON."""
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump({
            'hierarchy_tree': hierarchy['hierarchy_tree'],
            'all_rows_involved': hierarchy['all_rows_involved'],
            'stats': hierarchy['stats']
        }, f, indent=2, ensure_ascii=False)
    
    print(f"\n✓ Row hierarchy saved to: {output_path}")

def main():
    """Main function."""
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    data_dir = project_root / "data"
    
    xlsx_files = list(data_dir.glob("*.xlsx"))
    if not xlsx_files:
        print("Error: No XLSX files found")
        exit(1)
    
    xlsx_file = xlsx_files[0]
    print(f"Processing: {xlsx_file}")
    
    # Build row hierarchy
    hierarchy = build_formula_hierarchy(xlsx_file, column='K')
    
    # Save results
    output_file = data_dir / "row_hierarchy.json"
    save_row_hierarchy(hierarchy, output_file)
    
    print("\n" + "="*70)
    print("PHASE 1 COMPLETE: Row hierarchy built!")
    print("="*70)
    print(f"\nNext step: Run 'map_row_data_polars.py' to add labels and amounts")
    print(f"  This will map {len(hierarchy['all_rows_involved'])} row numbers to data")

if __name__ == "__main__":
    main()
