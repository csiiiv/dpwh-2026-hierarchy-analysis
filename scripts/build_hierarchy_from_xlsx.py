#!/usr/bin/env python3
"""
Build hierarchical JSON tree from XLSX data with formatting awareness.
Uses cell formatting (underline, italic) to identify titles, descriptions, and data rows.
Traverses rows and determines hierarchy based on which column the first data appears in.
"""

import sys
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

def is_bullet(value: str) -> bool:
    """
    Check if a value is a bullet/numbering marker (like 'a.', '1.0', '2.0', etc.).
    Bullets are typically short (1-3 characters) and used for list numbering.
    
    Args:
        value: Cell value to check
    
    Returns:
        True if the value appears to be a bullet marker
    """
    if not value:
        return False
    
    value = str(value).strip()
    
    # Bullets are typically very short (1-4 characters max)
    # Long values are definitely not bullets (like amounts: '18371150000.0')
    if len(value) > 4:
        return False
    
    # Single letter followed by period (like 'a.', 'b.')
    if len(value) == 2 and value[-1] == '.' and value[0].isalpha():
        return True
    
    # Number followed by period (like '1.', '2.', '10.')
    if value[-1] == '.' and value[:-1].isdigit():
        return True
    
    # Number with decimal (like '1.0', '2.0', '10.0') - but only if short
    if '.' in value and len(value) <= 4:
        parts = value.split('.')
        if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
            return True
    
    # Just a single character (like 'a', 'b', '1', '2')
    if len(value) == 1 and (value.isalpha() or value.isdigit()):
        return True
    
    return False

def get_cell_formatting(cell: Cell) -> Dict[str, bool]:
    """
    Get formatting information from a cell.
    
    Args:
        cell: OpenPyXL Cell object
    
    Returns:
        Dictionary with formatting flags
    """
    if cell is None:
        return {"underline": False, "italic": False, "bold": False}
    
    font = cell.font
    return {
        "underline": font.underline is not None and font.underline != "none",
        "italic": font.italic or False,
        "bold": font.bold or False,
    }

def find_hierarchy_info(row: List[Cell], start_col: int = 1, amount_col: int = 10) -> Optional[tuple]:
    """
    Find hierarchy level and value column in a row, considering formatting.
    If a bullet marker is found, the bullet's column determines hierarchy level,
    and the value is in the column immediately to the right.
    If no bullet is found, the first non-empty column is both level and value location.
    
    Args:
        row: List of Cell objects
        start_col: Column index to start searching from (default: 1)
        amount_col: Column index containing amounts (default: 10)
    
    Returns:
        Tuple of (hierarchy_column, value_column, is_description) or None if row is empty
        is_description: True if this row appears to be a description (italic, no amount)
    """
    # First, look for bullet markers
    for i in range(start_col, len(row)):
        cell = row[i] if i < len(row) else None
        if cell and cell.value:
            value = str(cell.value).strip()
            if is_bullet(value):
                # Found a bullet - hierarchy level is this column, value is next column
                value_col = i + 1
                if value_col < len(row) and row[value_col] and row[value_col].value:
                    value_cell = row[value_col]
                    # Check if it's a description (italic formatting)
                    formatting = get_cell_formatting(value_cell)
                    is_description = formatting["italic"]
                    return (i, value_col, is_description)
                # Bullet found but no value in next column - skip this row
                return None
    
    # No bullet found - find first non-empty column (it's both level and value)
    for i in range(start_col, len(row)):
        cell = row[i] if i < len(row) else None
        if cell and cell.value:
            value = str(cell.value).strip()
            # Make sure it's not a bullet (shouldn't happen, but just in case)
            if not is_bullet(value):
                # Check formatting
                formatting = get_cell_formatting(cell)
                is_description = formatting["italic"]
                return (i, i, is_description)
    
    return None

def parse_hierarchical_xlsx(xlsx_path: Path, value_column: int = 10, start_column: int = 1, row_range: tuple = None) -> List[Dict[str, Any]]:
    """
    Parse XLSX file and build hierarchical tree structure with formatting awareness.
    The hierarchy is determined by which column the first data appears in.
    Lower column numbers = higher in hierarchy (parent).
    Higher column numbers = lower in hierarchy (child).
    
    Descriptions (italic text, no amount) are skipped and not added to the hierarchy.
    
    Args:
        xlsx_path: Path to XLSX file
        value_column: Column index containing the value/amount (default: 10)
        start_column: Column index to start looking for data (default: 1)
        row_range: Optional tuple (start, end) to process only specific rows (1-indexed, inclusive)
    
    Returns:
        List of root nodes in the hierarchy
    """
    print(f"Reading XLSX file: {xlsx_path}")
    
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active
    
    root_nodes = []
    node_stack = []  # Stack to track current path in hierarchy (each element is a node)
    
    # Get all rows
    rows = list(sheet.iter_rows())
    
    for row_num, row in enumerate(rows, start=1):
        # Apply row range filter if specified
        if row_range:
            start_row, end_row = row_range
            if row_num < start_row or row_num > end_row:
                continue
        
        # Skip completely empty rows
        if not any(cell and cell.value and str(cell.value).strip() for cell in row):
            continue
        
        # Find hierarchy level and value column
        hierarchy_info = find_hierarchy_info(row, start_col=start_column, amount_col=value_column)
        
        if hierarchy_info is None:
            # Row only contains bullets without values or is empty, skip
            continue
        
        hierarchy_col, value_col, is_description = hierarchy_info
        
        # Get the data value from the value column
        if value_col >= len(row) or not row[value_col] or not row[value_col].value:
            continue
        
        data_value = str(row[value_col].value).strip()
        if not data_value:
            continue
        
        # Check if this is a description row (italic, no amount)
        # Descriptions should be skipped
        amount = None
        amount_cell = row[value_column] if value_column < len(row) else None
        
        if amount_cell and amount_cell.value:
            try:
                amount_str = str(amount_cell.value).strip().replace(',', '')
                if amount_str:
                    amount = float(amount_str)
            except (ValueError, AttributeError):
                pass
        
        # Handle description rows (italic formatting and no amount)
        # Attach them to the most recent node instead of creating a new node
        if is_description and amount is None:
            # This is a description row - attach it to the most recent node
            # The most recent node should be the one that this description describes
            if node_stack:
                # Add description to the most recent node
                if "description" not in node_stack[-1]:
                    node_stack[-1]["description"] = data_value
                else:
                    # If description already exists, append to it (in case description spans multiple rows)
                    node_stack[-1]["description"] += " " + data_value
            # Skip creating a new node for descriptions
            continue
        
        # Create node
        node = {
            "value": data_value,
            "amount": amount,
            "children": []
        }
        
        # Find the correct parent by going up the stack
        # We need to find the most recent node with a hierarchy column < current hierarchy column
        # This handles cases where levels skip (e.g., 0 -> 2, skipping 1)
        parent = None
        while node_stack:
            candidate = node_stack[-1]
            # Get the hierarchy column of the candidate
            candidate_col = candidate.get("_hierarchy_col", start_column)
            if candidate_col < hierarchy_col:
                # This is a valid parent
                parent = candidate
                break
            else:
                # This node is at same or deeper level, pop it
                node_stack.pop()
        
        # Store hierarchy column for next iteration
        node["_hierarchy_col"] = hierarchy_col
        
        # Add node to appropriate parent
        if parent is None:
            # Root level node
            root_nodes.append(node)
        else:
            # Child node - add to parent's children
            parent["children"].append(node)
        
        # Push current node to stack
        node_stack.append(node)
        
        # Progress indicator
        if row_num % 10000 == 0:
            print(f"  Processed {row_num} rows...", end='\r')
    
    print(f"\n✓ Processed {row_num} rows")
    print(f"✓ Found {len(root_nodes)} root nodes")
    
    # Clean up temporary _hierarchy_col fields
    def clean_node(n):
        if "_hierarchy_col" in n:
            del n["_hierarchy_col"]
        for child in n["children"]:
            clean_node(child)
    
    for node in root_nodes:
        clean_node(node)
    
    workbook.close()
    return root_nodes

def count_nodes(nodes: List[Dict[str, Any]]) -> Dict[str, int]:
    """Count nodes in the tree structure."""
    total = len(nodes)
    with_children = sum(1 for n in nodes if n["children"])
    with_amount = sum(1 for n in nodes if n["amount"] is not None)
    
    for node in nodes:
        child_counts = count_nodes(node["children"])
        total += child_counts["total"]
        with_children += child_counts["with_children"]
        with_amount += child_counts["with_amount"]
    
    return {
        "total": total,
        "with_children": with_children,
        "with_amount": with_amount
    }

def main():
    """Main function to build hierarchy."""
    # Get the script directory and project root
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    data_dir = project_root / "data"
    
    # Find XLSX file in data directory
    xlsx_files = list(data_dir.glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"Error: No XLSX files found in {data_dir}")
        sys.exit(1)
    
    # Use the first XLSX file found
    xlsx_file = xlsx_files[0]
    
    try:
        # Parse hierarchical structure
        # Value column is typically column K (index 10)
        # Start looking for data from column B (index 1)
        # For testing, you can specify a row range: row_range=(5, 104)
        hierarchy = parse_hierarchical_xlsx(xlsx_file, value_column=10, start_column=1, row_range=None)
        
        # Count nodes
        counts = count_nodes(hierarchy)
        print(f"\nTree Statistics:")
        print(f"  Total nodes: {counts['total']}")
        print(f"  Nodes with children: {counts['with_children']}")
        print(f"  Nodes with amounts: {counts['with_amount']}")
        
        # Output results
        output_file = data_dir / f"{xlsx_file.stem}_hierarchy.json"
        
        print(f"\nWriting hierarchy to: {output_file}")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(hierarchy, f, indent=2, ensure_ascii=False)
        
        print(f"✓ Hierarchy saved to {output_file}")
        
        # Print sample of first few levels
        print("\n" + "="*60)
        print("SAMPLE HIERARCHY (first 3 levels):")
        print("="*60)
        def print_sample(nodes, level=0, max_level=2, max_items=3):
            if level > max_level:
                return
            for i, node in enumerate(nodes[:max_items]):
                indent = "  " * level
                amount_str = f" (Amount: {node['amount']})" if node['amount'] else ""
                print(f"{indent}- Level {level}: {node['value'][:60]}{amount_str}")
                if node['children']:
                    print_sample(node['children'], level + 1, max_level, max_items)
            if len(nodes) > max_items:
                print(f"{'  ' * level}... ({len(nodes) - max_items} more)")
        
        print_sample(hierarchy)
        
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
